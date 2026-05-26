from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).resolve().parents[1] / "test_commission_statement_may_2026.xlsx"

headers = [
    "Loan ID",
    "Borrower Reference",
    "Lender",
    "Settlement Date",
    "Original Loan Amount",
    "Current Balance",
    "Trail Rate %",
    "Trail Commission",
    "Upfront Commission",
]

rows = [
    ["YLD-1001", "Client A", "Commonwealth Bank", "2022-08-15", 720000, 646250, 0.15, 80.78, 0],
    ["YLD-1002", "Client B", "Westpac", "2021-03-22", 540000, 421800, 0.16, 56.24, 0],
    ["YLD-1003", "Client C", "ANZ", "2024-01-12", 910000, 882500, 0.14, 102.96, 0],
    ["YLD-1004", "Client D", "NAB", "2020-11-03", 615000, 377400, 0.17, 53.47, 0],
    ["YLD-1005", "Client E", "Macquarie Bank", "2023-06-30", 825000, 794200, 0.15, 99.28, 0],
    ["YLD-1006", "Client F", "ING", "2019-09-18", 455000, 291750, 0.18, 43.76, 0],
    ["YLD-1007", "Client G", "Bankwest", "2025-02-07", 680000, 672100, 0.14, 78.41, 0],
    ["YLD-1008", "Client H", "Suncorp", "2022-12-01", 735000, 691350, 0.16, 92.18, 0],
    ["YLD-1009", "Client I", "Commonwealth Bank", "2021-07-29", 980000, 803900, 0.15, 100.49, 0],
    ["YLD-1010", "Client J", "Westpac", "2024-09-16", 510000, 498650, 0.14, 58.18, 0],
    ["YLD-1011", "Client K", "ANZ", "2023-04-04", 760000, 718300, 0.15, 89.79, 0],
    ["YLD-1012", "Client L", "NAB", "2020-02-25", 590000, 341100, 0.17, 48.32, 0],
]

wb = Workbook()
ws = wb.active
ws.title = "Commission Statement"

ws["A1"] = "Yield Test Aggregator - Commission Statement"
ws["A2"] = "Statement Period"
ws["B2"] = "2026-05"
ws["A3"] = "Note"
ws["B3"] = "Synthetic test data for upload validation"

header_row = 5
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0E7490")
    cell.alignment = Alignment(horizontal="center")

for row_index, row in enumerate(rows, start=header_row + 1):
    for col_index, value in enumerate(row, start=1):
        ws.cell(row=row_index, column=col_index, value=value)

currency_cols = [5, 6, 8, 9]
for col in currency_cols:
    for row in range(header_row + 1, header_row + 1 + len(rows)):
        ws.cell(row=row, column=col).number_format = '$#,##0.00'

for row in range(header_row + 1, header_row + 1 + len(rows)):
    ws.cell(row=row, column=7).number_format = '0.00'
    ws.cell(row=row, column=4).number_format = 'yyyy-mm-dd'

thin = Side(style="thin", color="CBD5E1")
for row in ws.iter_rows(min_row=header_row, max_row=header_row + len(rows), min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(vertical="center")

ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A{header_row}:I{header_row + len(rows)}"

widths = [14, 20, 24, 16, 20, 18, 14, 18, 20]
for index, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(index)].width = width

ws["A1"].font = Font(bold=True, size=14, color="0F172A")
ws["A2"].font = Font(bold=True)
ws["A3"].font = Font(bold=True)

wb.save(OUTPUT)
print(OUTPUT)
