import openpyxl
import xlrd
import io
from typing import List, Dict, Any, Union

def _find_header_row_index(sheet_data: List[List[Any]]) -> int:
    """
    Heuristically find the actual table header row in a spreadsheet.
    Account for summary rows at the top.
    """
    if not sheet_data:
        return 0
        
    best_row_idx = 0
    best_score = -1
    
    keywords = {
        'trail', 'amount', 'date', 'loan', 'balance', 'name', 
        'borrower', 'lender', 'commission', 'settlement', 'client'
    }
    
    # Check max first 50 rows
    for i, row in enumerate(sheet_data[:50]):
        score = 0
        non_empty_count = 0
        
        for cell in row:
            if cell is not None and str(cell).strip() != "":
                non_empty_count += 1
                val = str(cell).lower().strip()
                if any(k in val for k in keywords):
                    score += 2
                    
        total_score = score + non_empty_count
        if total_score > best_score:
            best_score = total_score
            best_row_idx = i
            
    return best_row_idx

def _extract_data_from_sheet(sheet_data: List[List[Any]]) -> List[Dict[str, Any]]:
    if not sheet_data:
        return []
        
    header_idx = _find_header_row_index(sheet_data)
    headers = []
    
    raw_headers = sheet_data[header_idx]
    for col_idx, h in enumerate(raw_headers):
        if h is not None and str(h).strip() != "":
            headers.append(str(h).strip().replace('\n', ' '))
        else:
            headers.append(f"Column_{col_idx}")
            
    data_dicts = []
    
    # Iterate dynamically through parsed rows
    for row in sheet_data[header_idx + 1:]:
        # Skip fully empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
            
        row_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = cell
                
        data_dicts.append(row_dict)
        
    return data_dicts

def _parse_xlsx(file_path_or_bytes: Union[str, bytes]) -> List[Dict[str, Any]]:
    """Uses openpyxl to parse standard .xlsx files, returning a list of dicts."""
    if isinstance(file_path_or_bytes, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(file_path_or_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True)
        
    best_sheet = None
    max_rows = -1
    
    # Identify data sheet based on which sheet has the most data rows
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        if sheet.max_row > max_rows:
            max_rows = sheet.max_row
            best_sheet = sheet
            
    # Extract sheet data to standard list format
    sheet_data = []
    if best_sheet:
        for row in best_sheet.iter_rows(values_only=True):
            sheet_data.append(list(row))
            
    return _extract_data_from_sheet(sheet_data)

def _parse_xls(file_path_or_bytes: Union[str, bytes]) -> List[Dict[str, Any]]:
    """Uses xlrd to parse legacy .xls files, returning a list of dicts."""
    if isinstance(file_path_or_bytes, bytes):
        wb = xlrd.open_workbook(file_contents=file_path_or_bytes)
    else:
        wb = xlrd.open_workbook(file_path_or_bytes)
        
    best_sheet = None
    max_rows = -1
    
    # Identify data sheet
    for i in range(wb.nsheets):
        sheet = wb.sheet_by_index(i)
        if sheet.nrows > max_rows:
            max_rows = sheet.nrows
            best_sheet = sheet
            
    sheet_data = []
    if best_sheet:
        for row_idx in range(best_sheet.nrows):
            sheet_data.append(best_sheet.row_values(row_idx))
            
    return _extract_data_from_sheet(sheet_data)

def parse_excel(file_path_or_bytes: Union[str, bytes], file_name: str) -> List[Dict[str, Any]]:
    """
    Main entry point for parsing Excel commission statements.
    Automatically identifies legacy vs modern excel format by filename.
    """
    is_legacy = file_name.lower().endswith('.xls')
    
    if is_legacy:
        return _parse_xls(file_path_or_bytes)
    else:
        return _parse_xlsx(file_path_or_bytes)
