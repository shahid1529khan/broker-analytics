from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]

headers = [
    "Loan ID",
    "Borrower Reference",
    "Lender",
    "Settlement Date",
    "Original Loan Amount",
    "Current Balance",
    "Trail Rate %",
    "Trail Commission",
    "Upfront Commission",
]

periods = {
    "2026-02": [
        ["YLD-1001", "Client A", "Commonwealth Bank", "2022-08-15", 720000, 659800, 0.15, 82.48, 0],
        ["YLD-1002", "Client B", "Westpac", "2021-03-22", 540000, 432900, 0.16, 57.72, 0],
        ["YLD-1003", "Client C", "ANZ", "2024-01-12", 910000, 895300, 0.14, 104.45, 0],
        ["YLD-1004", "Client D", "NAB", "2020-11-03", 615000, 390600, 0.17, 55.34, 0],
        ["YLD-1005", "Client E", "Macquarie Bank", "2023-06-30", 825000, 804900, 0.15, 100.61, 0],
        ["YLD-1006", "Client F", "ING", "2019-09-18", 455000, 302400, 0.18, 45.36, 0],
        ["YLD-1008", "Client H", "Suncorp", "2022-12-01", 735000, 701800, 0.16, 93.57, 0],
        ["YLD-1009", "Client I", "Commonwealth Bank", "2021-07-29", 980000, 819600, 0.15, 102.45, 0],
        ["YLD-1011", "Client K", "ANZ", "2023-04-04", 760000, 730200, 0.15, 91.28, 0],
        ["YLD-1012", "Client L", "NAB", "2020-02-25", 590000, 353800, 0.17, 50.12, 0],
    ],
    "2026-03": [
        ["YLD-1001", "Client A", "Commonwealth Bank", "2022-08-15", 720000, 655100, 0.15, 81.89, 0],
        ["YLD-1002", "Client B", "Westpac", "2021-03-22", 540000, 429050, 0.16, 57.21, 0],
        ["YLD-1003", "Client C", "ANZ", "2024-01-12", 910000, 891400, 0.14, 104.0, 0],
        ["YLD-1004", "Client D", "NAB", "2020-11-03", 615000, 386200, 0.17, 54.71, 0],
        ["YLD-1005", "Client E", "Macquarie Bank", "2023-06-30", 825000, 801100, 0.15, 100.14, 0],
        ["YLD-1006", "Client F", "ING", "2019-09-18", 455000, 298700, 0.18, 44.81, 0],
        ["YLD-1007", "Client G", "Bankwest", "2025-02-07", 680000, 676500, 0.14, 78.93, 0],
        ["YLD-1008", "Client H", "Suncorp", "2022-12-01", 735000, 698250, 0.16, 93.1, 0],
        ["YLD-1009", "Client I", "Commonwealth Bank", "2021-07-29", 980000, 813800, 0.15, 101.73, 0],
        ["YLD-1011", "Client K", "ANZ", "2023-04-04", 760000, 725500, 0.15, 90.69, 0],
        ["YLD-1012", "Client L", "NAB", "2020-02-25", 590000, 349200, 0.17, 49.47, 0],
    ],
    "2026-04": [
        ["YLD-1001", "Client A", "Commonwealth Bank", "2022-08-15", 720000, 650700, 0.15, 81.34, 0],
        ["YLD-1002", "Client B", "Westpac", "2021-03-22", 540000, 425600, 0.16, 56.75, 0],
        ["YLD-1003", "Client C", "ANZ", "2024-01-12", 910000, 887100, 0.14, 103.5, 0],
        ["YLD-1004", "Client D", "NAB", "2020-11-03", 615000, 381600, 0.17, 54.06, 0],
        ["YLD-1005", "Client E", "Macquarie Bank", "2023-06-30", 825000, 797400, 0.15, 99.68, 0],
        ["YLD-1006", "Client F", "ING", "2019-09-18", 455000, 295000, 0.18, 44.25, 0],
        ["YLD-1007", "Client G", "Bankwest", "2025-02-07", 680000, 674300, 0.14, 78.67, 0],
        ["YLD-1008", "Client H", "Suncorp", "2022-12-01", 735000, 694800, 0.16, 92.64, 0],
        ["YLD-1009", "Client I", "Commonwealth Bank", "2021-07-29", 980000, 808600, 0.15, 101.08, 0],
        ["YLD-1010", "Client J", "Westpac", "2024-09-16", 510000, 503800, 0.14, 58.78, 0],
        ["YLD-1011", "Client K", "ANZ", "2023-04-04", 760000, 721600, 0.15, 90.2, 0],
        ["YLD-1012", "Client L", "NAB", "2020-02-25", 590000, 345000, 0.17, 48.88, 0],
    ],
}


def save_statement(period: str, rows: list[list[object]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Commission Statement"

    ws["A1"] = "Yield Test Aggregator - Commission Statement"
    ws["A2"] = "Statement Period"
    ws["B2"] = period
    ws["A3"] = "Note"
    ws["B3"] = "Synthetic test data for upload validation"

    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0E7490")
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=header_row + 1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    for col in [5, 6, 8, 9]:
        for row in range(header_row + 1, header_row + 1 + len(rows)):
            ws.cell(row=row, column=col).number_format = '$#,##0.00'

    for row in range(header_row + 1, header_row + 1 + len(rows)):
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=4).number_format = "yyyy-mm-dd"

    thin = Side(style="thin", color="CBD5E1")
    for row in ws.iter_rows(min_row=header_row, max_row=header_row + len(rows), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:I{header_row + len(rows)}"

    widths = [14, 20, 24, 16, 20, 18, 14, 18, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws["A1"].font = Font(bold=True, size=14, color="0F172A")
    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(bold=True)

    output = ROOT / f"test_commission_statement_{period}.xlsx"
    wb.save(output)
    return output


for period, rows in periods.items():
    print(save_statement(period, rows))
